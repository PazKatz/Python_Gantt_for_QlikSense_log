import os #read all log files with loop for
import re
# import datetime  # foramt hh:mm
import xlsxwriter as xls  # output excel file
from openpyxl import Workbook , load_workbook
import plotly.express as px
from pathlib import Path
import pandas as pd
import plotly.express as px


workbook = xls.Workbook("Gant_Task_Qlik.xlsx")  # Name Excel file
worksheet = workbook.add_worksheet("Tasks_timeline")  # Name Excel TAB
wb=load_workbook('Gant_Task_Qlik.xlsx')
worksheet.write(0,0,"Task num")
worksheet.write(0,1,"Start")
worksheet.write(0,2,"End")
worksheet.write(0,3,"Duration")
worksheet.write(0,4,"file name with path")
################################################
os # read all log files with loop for
re # search lines with patterns strings + readlines

folderpath = input('Please insert a folder path of log files from Qlik Sense\n')  # r"C:\Users\97250\Desktop\logs"
filepaths = [os.path.join(folderpath, name) for name in os.listdir(folderpath)]
all_files = []

# Insert start+end time from log text file of qlik sense :
index_xls_r = 1  # row index promoted
col_xls_start = 1  # column static num
col_xls_end = 2  # column static num
col_xls_duration = 3  # column static num
for path in filepaths: # building array of all log text files
    with open(path, "rt",encoding='utf-8') as my_log:
      a = [re.search(r'\d{8}[A-Z]{1}\d{6}', line)[0] for line in my_log.readlines() if 'Execution started' in line or 'Execution finished' in line]

####################################################
# Create Excel File:
    start = a[0][10:12]+':'+a[0][12:14]  # +':'+a[0][14:15]
    end = a[1][10:12]+':'+a[1][12:14]  # +':'+a[1][14:15]

    import datetime

    start_time = start  # datetime.datetime.strptime(start, '%H:%M').time() # datetime format, problem with plot
    end_time = end  # datetime.datetime.strptime(end, '%H:%M').time() # datetime format, problem with plot

    # duration calc:
    start_time_for_duration_calc = datetime.datetime.strptime(start, '%H:%M').time()  # duration calculation
    end_time_for_duration_calc = datetime.datetime.strptime(end, '%H:%M').time()  # duration calculation

    from datetime import datetime, date  # duration calc

    duration = str((datetime.combine(date.today(), end_time_for_duration_calc) - datetime.combine(date.today(), start_time_for_duration_calc)))
    # output excel duration:
    format_hh_mm = workbook.add_format({'num_format': 'hh:mm'})  # excel cell output hh:mm

    worksheet.write(index_xls_r, col_xls_start, start_time)  # start
    worksheet.write(index_xls_r, col_xls_end, end_time)  # end
    worksheet.write(index_xls_r, 0, index_xls_r)  # task name as number
    worksheet.write(index_xls_r, col_xls_duration, duration)  # duration
    worksheet.write(index_xls_r, 4, path)  # task name
    index_xls_r += 1
#################################################
#creat plot figure - Gantt
EXCEL_FILE = Path.cwd() / "Gant_Task_Qlik.xlsx"

# Read Dataframe from Excel file
df = pd.read_excel(EXCEL_FILE)
Tasks = df["Task num"]
Start = df["Start"]
Finish = df["End"]
Duration = df["Duration"]

print(df[["Task num", 'Start','End','Duration']])
df['Start'] = pd.to_datetime(df['Start'].str.split().str[-1])  # becomes date y-m-d with hh:mm
df['End'] = pd.to_datetime(df['End'].str.split().str[-1])

#import plotly.express as px
fig = px.timeline(df, x_start=df["Start"], x_end=df['End'], y=df["Task num"]  # , color="Task_Name"
                 )


fig = px.timeline(df, x_start="Start", x_end="End", y="Task num" #, color="Task_Name"
                 )

fig.update_layout(xaxis=dict(
                      title='Timestamp',
                      tickformat = '%H:%M',
                  ))
fig.show()

workbook.close()
